"""
QRコード作成ツール
"""

import customtkinter
import tkinter
import os
import qrcode
import threading
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from tkinter import filedialog, messagebox
from PIL import Image


customtkinter.set_default_color_theme("green")


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # メンバー変数の設定
        self.fonts = ("", 30)
        # フォームサイズ設定
        width = 600
        height = 350
        self.geometry(f"{width}x{height}")
        self.minsize(width, height)
        self.maxsize(width, height)
        self.title("QRコード作成ツール")

        # パーツのセット
        self.set_up()

    def set_up(self):
        self.tabview = customtkinter.CTkTabview(
            master=self, width=550, height=320, anchor="nw"
        )
        self.tabview.add("単体作成")
        self.tabview.add("一括作成")
        self.tabview.set("一括作成")
        self.tabview.place(x=10, y=10)

        # 単体作成タブ
        self.tabview1_label1 = customtkinter.CTkLabel(
            master=self.tabview.tab("単体作成"), text="【URL】"
        )
        self.tabview1_label1.place(x=10, y=10)

        self.tabview_entry1 = customtkinter.CTkEntry(
            master=self.tabview.tab("単体作成"),
            width=500,
        )
        self.tabview_entry1.place(x=10, y=40)

        self.tabview1_label1 = customtkinter.CTkLabel(
            master=self.tabview.tab("単体作成"), text="【ファイル名(拡張子無し)】"
        )
        self.tabview1_label1.place(x=10, y=80)

        self.tabview_entry2 = customtkinter.CTkEntry(
            master=self.tabview.tab("単体作成"),
            width=500,
        )
        self.tabview_entry2.place(x=10, y=110)

        self.tabview1_label2 = customtkinter.CTkLabel(
            master=self.tabview.tab("単体作成"),
            text="【ファイル形式】選択しない場合両方作成します。",
        )
        self.tabview1_label2.place(x=10, y=150)

        self.mode = tkinter.StringVar()
        self.tabview1_radiobutton1 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("単体作成"),
            variable=self.mode,
            text="JPEG形式",
            value="jpg",
        )
        self.tabview1_radiobutton1.place(x=10, y=180)

        self.tabview1_radiobutton2 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("単体作成"),
            variable=self.mode,
            text="EPS形式",
            value="eps",
        )
        self.tabview1_radiobutton2.place(x=10, y=210)

        self.tabview1_radiobutton3 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("単体作成"),
            variable=self.mode,
            text="EPSとJPEG両方",
            value="both",
        )
        self.tabview1_radiobutton3.place(x=10, y=240)

        self.tabview1_button1 = customtkinter.CTkButton(
            master=self.tabview.tab("単体作成"),
            text="作成",
            width=80,
            command=self.make_qr_one,
        )
        self.tabview1_button1.place(x=450, y=240)

        # 一括作成タブ
        self.tabview2_label1 = customtkinter.CTkLabel(
            master=self.tabview.tab("一括作成"), text="Excelファイル："
        )
        self.tabview2_label1.place(x=10, y=10)

        self.tabview2_entry1 = customtkinter.CTkEntry(
            master=self.tabview.tab("一括作成"),
            width=450,
        )
        self.tabview2_entry1.place(x=10, y=40)

        self.tabview2_button1 = customtkinter.CTkButton(
            master=self.tabview.tab("一括作成"),
            text="選択",
            width=50,
            command=self.select_folder,
        )
        self.tabview2_button1.place(x=470, y=40)

        self.tabview2_label2 = customtkinter.CTkLabel(
            master=self.tabview.tab("一括作成"),
            text="【ファイル形式】選択しない場合両方作成します。",
        )
        self.tabview2_label2.place(x=10, y=150)

        self.all_mode = tkinter.StringVar()
        self.tabview2_radiobutton1 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("一括作成"),
            variable=self.all_mode,
            text="JPEG形式",
            value="jpg",
        )
        self.tabview2_radiobutton1.place(x=10, y=180)

        self.tabview2_radiobutton2 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("一括作成"),
            variable=self.all_mode,
            text="EPS形式",
            value="eps",
        )
        self.tabview2_radiobutton2.place(x=10, y=210)

        self.tabview2_radiobutton3 = customtkinter.CTkRadioButton(
            master=self.tabview.tab("一括作成"),
            variable=self.all_mode,
            text="EPSとJPEG両方",
            value="both",
        )
        self.tabview2_radiobutton3.place(x=10, y=240)

        self.tabview2_button1 = customtkinter.CTkButton(
            master=self.tabview.tab("一括作成"),
            text="作成",
            command=self.qr_thread,
            width=80,
        )
        self.tabview2_button1.place(x=450, y=240)

    # エクスプローラー表示
    def select_folder(self):
        folder_path = filedialog.askopenfilename(
            title="Excelファイルの選択",
        )
        self.tabview2_entry1.delete(0, tkinter.END)
        self.tabview2_entry1.insert(0, folder_path)

    def qr_thread(self):
        make_thread = threading.Thread(target=self.make_qr_all)
        try:
            make_thread.start()

        except:
            messagebox.showerror(
                title="エラー",
                message="QRコードの作成に失敗しました。\n再度実行してください。",
            )

    # まとめて作成処理
    def make_qr_all(self):
        file_path = self.tabview2_entry1.get()
        mode = self.all_mode.get()

        if file_path == "":
            messagebox.showerror(
                title="エラー",
                message="入力していない項目があります。\n再度確認してください。",
            )
        else:
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
            except (FileNotFoundError, InvalidFileException):
                messagebox.showerror(
                    title="エラー",
                    message="ファイルを読み込めませんでした。パスと形式を確認してください。",
                )
                return
            sheet = workbook.worksheets[0]
            last_row = sheet.max_row
            flag = True

            for i in range(2, last_row + 1):
                if sheet.cell(i, 1).value == None or sheet.cell(i, 2).value == None:
                    flag = False
                    break

            if flag:
                # 作成
                for i in range(2, last_row + 1):
                    self.make_qr(mode, sheet.cell(i, 2).value, sheet.cell(i, 1).value)
                messagebox.showinfo("完了", "QRコードが生成されました。")
                self.tabview2_entry1.delete(0, len(self.tabview2_entry1.get()) + 1)

            else:
                messagebox.showerror(
                    title="エラー",
                    message="データに不備があります。\n再度確認してください。",
                )

    # 1つだけ作成処理
    def make_qr_one(self):
        url = self.tabview_entry1.get()
        name = self.tabview_entry2.get()
        mode = self.mode.get()

        if url == "" or name == "":
            messagebox.showerror(
                title="エラー",
                message="入力していない項目があります。再度確認してください。",
            )

        else:
            try:
                self.make_qr(mode, url, name)
                messagebox.showinfo("完了", "QRコードが生成されました。")
                self.tabview_entry1.delete(0, len(self.tabview_entry1.get()) + 1)
                self.tabview_entry2.delete(0, len(self.tabview_entry2.get()) + 1)
            except:
                messagebox.showerror(
                    title="エラー", message="QRコードの作成に失敗しました"
                )

    # QRコード作成
    def make_qr(self, mode, url, file_name):
        # 出力先の作成
        data_dir = os.path.expanduser("~/Desktop") + "/QR-Code/"
        os.makedirs(data_dir, exist_ok=True)
        # QRデータ作成
        img = qrcode.make(url)
        file_path = data_dir + file_name + ".jpg"
        img.save(file_path)

        if mode == "jpg":
            pass
        elif mode == "eps":
            img = Image.open(file_path)
            fig = img.convert("RGB")
            fig.save(data_dir + file_name.split(".", 1)[0] + ".eps", lossless=True)
            os.remove(file_path)
        else:
            img = Image.open(file_path)
            fig = img.convert("RGB")
            fig.save(data_dir + file_name.split(".", 1)[0] + ".eps", lossless=True)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
